# -*- coding: utf-8 -*-
"""
P&L-агент Градекс KZ — мультизагрузка месяцев в один отчёт.
"""
import io, os, re, tempfile
from collections import defaultdict
import streamlit as st
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import pnl_agent as eng

st.set_page_config(page_title="P&L-агент Градекс KZ", page_icon="📊", layout="wide")
MONTHS=['Январь','Февраль','Март','Апрель','Май','Июнь','Июль','Август','Сентябрь','Октябрь','Ноябрь','Декабрь']
MONTH_COL={m:get_column_letter(3+i) for i,m in enumerate(MONTHS)}
MONTH_IDX={m:i for i,m in enumerate(MONTHS)}
STEMS=[('декабр','Декабрь'),('ноябр','Ноябрь'),('октябр','Октябрь'),('сентябр','Сентябрь'),
       ('август','Август'),('июл','Июль'),('июн','Июнь'),('апрел','Апрель'),('март','Март'),
       ('феврал','Февраль'),('январ','Январь'),('мая','Май'),('май','Май')]
NUMS={'01':'Январь','02':'Февраль','03':'Март','04':'Апрель','05':'Май','06':'Июнь',
      '07':'Июль','08':'Август','09':'Сентябрь','10':'Октябрь','11':'Ноябрь','12':'Декабрь'}

def guess_month(fn):
    s=fn.lower()
    for stem,m in STEMS:
        if stem in s: return m
    for num,m in NUMS.items():
        if re.search(rf'[_\-\.\s]{num}[_\-\.\s]', s) or s.startswith(num): return m
    return None

st.title("📊 P&L-агент Градекс KZ")
st.caption("Загрузите одну или несколько выгрузок 1С — агент разнесёт всё по дивизионам "
           "ТШО / КПО / Бетон / АУП и соберёт месяцы в один P&L по шаблону.")

with st.sidebar:
    st.header("Настройки")
    threshold=st.number_input("Порог новой статьи, ₸", value=300_000, step=50_000,
        help="Новая статья ≥ порога → отдельная строка; ниже → «Прочие».")
    eng.THRESHOLD=int(threshold)
    st.divider()
    st.subheader("Накопление")
    base_up=st.file_uploader("Предыдущий отчёт (.xlsx)", type=['xlsx'], key='base',
        help="Чтобы добавить новые месяцы к уже заполненным — загрузите сюда прошлый отчёт. "
             "Пусто — начать с чистого шаблона.")
    with st.expander("Заменить мастер-шаблон (редко)"):
        tmpl_up=st.file_uploader("Чистый шаблон P&L (.xlsx)", type=['xlsx'], key='tmpl')

st.subheader("Выгрузки 1С")
data_ups=st.file_uploader("Перетащите файлы (можно несколько — за разные месяцы)",
    type=['xlsx'], accept_multiple_files=True, key='data')

assignments=[]
if data_ups:
    st.caption("Проверьте месяц для каждого файла (определяется по названию, можно поменять):")
    used=set()
    for i,uf in enumerate(data_ups):
        c1,c2=st.columns([3,2])
        c1.write(f"📄 {uf.name}")
        g=guess_month(uf.name)
        default=MONTHS.index(g) if g else 0
        m=c2.selectbox("Месяц", MONTHS, index=default, key=f"m_{i}", label_visibility="collapsed")
        assignments.append((uf,m))
        if m in used: st.warning(f"Месяц «{m}» выбран более одного раза — последний файл перезапишет предыдущий.")
        used.add(m)

def resolve_base():
    if base_up is not None: return base_up.getvalue(),'предыдущий отчёт'
    if tmpl_up is not None: return tmpl_up.getvalue(),'загруженный шаблон'
    if os.path.exists('template.xlsx'): return open('template.xlsx','rb').read(),'чистый вшитый шаблон'
    return None,None


def restore_formulas(work_path, tmpl_bytes):
    """Накладывает формульный каркас из шаблона на рабочий файл (значения строк сохраняются)."""
    import io as _io
    tw=openpyxl.load_workbook(_io.BytesIO(tmpl_bytes))
    ww=openpyxl.load_workbook(work_path)
    for sn in tw.sheetnames:
        if sn not in ww.sheetnames: continue
        ts=tw[sn]; ws=ww[sn]
        for row in ts.iter_rows():
            for c in row:
                if isinstance(c.value,str) and c.value.startswith('='):
                    ws[c.coordinate].value=c.value
    ww.save(work_path)

def harden_diverr(wb):
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v=c.value
                if isinstance(v,str) and v.startswith('=') and 'IFERROR' not in v.upper():
                    if re.search(r'/[A-Z]{1,2}(5|7)\b', v[1:]):
                        c.value=f'=IFERROR({v[1:]},"")'

def build_journal(wb, rows):
    if 'Журнал сопоставления' in wb.sheetnames: del wb['Журнал сопоставления']
    js=wb.create_sheet('Журнал сопоставления')
    js.append(['Месяц','Счёт 1С','Статья 1С','Дивизион','Тип','Лист отчёта','Строка отчёта','Сумма ₸'])
    thin=Side(style='thin',color='D0D0D0'); bd=Border(thin,thin,thin,thin)
    for c in js[1]:
        c.font=Font(bold=True,color='FFFFFF',name='Arial',size=10); c.fill=PatternFill('solid',fgColor='2F5496')
        c.alignment=Alignment(horizontal='center',wrap_text=True); c.border=bd
    rows=sorted(rows,key=lambda x:(MONTH_IDX.get(x[0],99),x[5],str(x[6])))
    for r in rows: js.append([r[0],r[1],r[2],r[3],r[4],r[5],str(r[6]),round(float(r[7]),2)])
    for i,w in enumerate([10,9,44,9,11,18,40,15],1): js.column_dimensions[get_column_letter(i)].width=w
    last=js.max_row
    for rr in range(2,last+1):
        for cc in range(1,9):
            cell=js.cell(row=rr,column=cc); cell.font=Font(name='Arial',size=10); cell.border=bd
            if cc==8: cell.number_format='#,##0;(#,##0);-'; cell.alignment=Alignment(horizontal='right')
    js.freeze_panes='A2'
    js.cell(row=last+1,column=7,value='ИТОГО').font=Font(bold=True,name='Arial')
    t=js.cell(row=last+1,column=8,value=f'=SUM(H2:H{last})'); t.font=Font(bold=True,name='Arial'); t.number_format='#,##0'

run=st.button("▶️ Сформировать отчёт", type="primary", disabled=(not assignments))

if run:
    base_bytes,base_kind=resolve_base()
    if base_bytes is None:
        st.error("Не найдена основа отчёта."); st.stop()
    st.caption(f"Основа: {base_kind}. Месяцев загружено: {len(assignments)}.")
    all_log=[]; per_month=[]
    with tempfile.TemporaryDirectory() as td:
        cur=os.path.join(td,'cur.xlsx'); open(cur,'wb').write(base_bytes)
        # восстановить формульный каркас (если основа — ранее «вшитый» отчёт)
        try:
            tmpl_for_restore=open('template.xlsx','rb').read() if os.path.exists('template.xlsx') else base_bytes
            restore_formulas(cur, tmpl_for_restore)
        except Exception:
            pass
        prog=st.progress(0.0)
        for i,(uf,month) in enumerate(sorted(assignments,key=lambda x:MONTH_IDX[x[1]])):
            dp=os.path.join(td,f'd{i}.xlsx'); op=os.path.join(td,f'o{i}.xlsx')
            open(dp,'wb').write(uf.getvalue())
            try:
                leaves=eng.parse_1c(dp)
            except Exception as ex:
                st.error(f"«{uf.name}»: не удалось разобрать ({ex})"); st.stop()
            acc,flags,log=eng.fill_report(cur,leaves,op,month_col=MONTH_COL[month])
            cur=op
            for r in log: all_log.append((month,)+tuple(r))
            ref=defaultdict(float); inc=defaultdict(float)
            for (a,it,k,d,v) in leaves: (ref if k=='expense' else inc)[d]+=v
            per_month.append({'Месяц':month,'Файл':uf.name,'Доходы':sum(inc.values()),
                              'Расходы':sum(ref.values()),'Новых статей':len(flags)})
            prog.progress((i+1)/len(assignments))
        wb=openpyxl.load_workbook(cur); harden_diverr(wb); build_journal(wb, all_log)
        wb.calculation.fullCalcOnLoad=True
        final=os.path.join(td,'final.xlsx'); wb.save(final)
        try:
            import recalc_native; nset=recalc_native.recalc_native(final)
        except Exception as _e:
            nset=0; st.warning(f'Пересчёт значений не выполнен ({_e}); формулы посчитаются при открытии в Excel.')
        buf=io.BytesIO(open(final,'rb').read()); buf.seek(0)

    st.success(f"Готово. В отчёт собрано месяцев: {len(assignments)}.")
    pm=pd.DataFrame(per_month)
    pm_disp=pm.copy()
    for col in ['Доходы','Расходы']: pm_disp[col]=pm_disp[col].map(lambda x:f"{x:,.0f}")
    st.subheader("Сводка по месяцам")
    st.dataframe(pm_disp, hide_index=True, use_container_width=True)
    tot_inc=pm['Доходы'].sum(); tot_exp=pm['Расходы'].sum()
    a,b,c=st.columns(3)
    a.metric("Доходы всего, ₸", f"{tot_inc:,.0f}")
    b.metric("Расходы всего, ₸", f"{tot_exp:,.0f}")
    c.metric("Результат до КПН, ₸", f"{tot_inc-tot_exp:,.0f}")

    months_str="_".join(sorted({m for _,m in assignments}, key=lambda x:MONTH_IDX[x]))
    st.download_button("⬇️ Скачать заполненный P&L (.xlsx)", buf,
        file_name=f"GRD_P&L_{months_str}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
    st.info("Сохраните файл и в следующий раз загрузите его в «Предыдущий отчёт» — месяцы будут накапливаться. "
            "Файл пересчитывает формулы при открытии. Открывайте в Excel или Google Таблицах "
            "(не в быстром превью) — тогда все итоги посчитаются и у вас, и у получателя.")
else:
    st.info("Загрузите одну или несколько выгрузок 1С, проверьте месяцы и нажмите «Сформировать отчёт».")
