# -*- coding: utf-8 -*-
"""P&L агент Градекс KZ — ядро v2."""
import openpyxl
from collections import defaultdict

MONTH_COL='C'; THRESHOLD=300_000

def detect_div(name):
    n=(name or '').lower()
    if 'ауп' in n: return 'АУП'
    if 'кпо' in n: return 'КПО'
    if 'тшо' in n: return 'ТШО'
    if 'бетон' in n: return 'Бетон'
    if 'номенклатурная группа' in n: return 'Бетон'
    return None

def parse_1c(path):
    """ОСВ/Анализ счёта. В P&L только счета 6xxx/7xxx. Берётся одна сторона:
    расход=Об.Дебет (col5), доход=Об.Кредит (col6). Подразделение — на любом уровне >=2."""
    wb=openpyxl.load_workbook(path); ws=wb.worksheets[0]
    def lvl(r):
        rd=ws.row_dimensions[r]; return rd.outline_level if rd else 0
    def is_pl(code):
        c=str(code).strip(); return c[:1] in ('6','7') and c[:4].isdigit()
    leaves=[]; acct=None; pl=False; item=None; item_val=None; item_kind=None; item_got=False; stack=[]
    def flush():
        if pl and item is not None and item_val and not item_got:
            leaves.append((acct,item,item_kind,None,float(item_val)))
    for r in range(1,ws.max_row+1):
        a=ws.cell(row=r,column=1).value
        if a is None: continue
        o=lvl(r)
        if a=='Итого' and o==0: continue
        e=ws.cell(row=r,column=5).value; f=ws.cell(row=r,column=6).value
        if o<=1:
            flush(); acct=str(a).split(',')[0].strip(); pl=is_pl(acct)
            if acct=='5610':
                v=ws.cell(row=r,column=6).value
                if v is not None: leaves.append(('5610','__CONTROL__','control',None,float(v)))
            elif acct=='7710':
                v=ws.cell(row=r,column=5).value
                if v is not None: leaves.append(('7710','__CONTROL__','control',None,float(v)))
            item=None; item_val=None; item_got=False; stack=[]; continue
        if not pl: continue
        if a=='<...>':
            stack=[(lv,dv) for (lv,dv) in stack if lv<o]; continue
        div=detect_div(a)
        if div and item is not None:
            stack=[(lv,dv) for (lv,dv) in stack if lv<o]
            anc=set(dv for (_,dv) in stack); val=e if e is not None else f
            if div not in anc and val:
                kind='income' if (f is not None and e is None) else 'expense'
                leaves.append((acct,item,kind,div,float(val))); item_got=True
            stack.append((o,div)); continue
        if o==2:
            flush(); item=a; stack=[]
            item_val=e if e is not None else f
            item_kind='income' if (f is not None and e is None) else 'expense'
            item_got=False; continue
        stack=[(lv,dv) for (lv,dv) in stack if lv<o]; stack.append((o,None))
    flush()
    return leaves

def norm(s): return (s or '').strip().lower()

INCOME_MAP={
 ('6010',norm('1доход от реализации товаров')):('Бетон','INC_CONCRETE'),
 ('6010',norm('2Доход от сдачи в аренду')):('Бетон','INC_RENT'),
 ('6010',norm('Доход от реализации продукции и оказания услуг')):('ТШО','INC_SALES'),
 ('6110',norm('Вознаграждение по депозиту')):('ТШО','INC_FIN'),
 ('6230',norm('Доход от гос субсидий')):('Бетон','INC_SUBSIDY'),
}
OPEX_CONCEPT={
 'gps':'GPS','амортизация фа':'AMORT','аренда офиса':'RENT_OFFICE','аренда спецтехники':'RENT_TECH',
 'аренда транспорта':'RENT_TECH','база расходы по содержанию':'BASE_MAINT','гсм':'FUEL','зарплата':'SALARY',
 'интернет':'INTERNET','услуги связи':'INTERNET','канц товары':'OTHER','ком.услуги':'UTILITIES',
 'электроэнергия':'UTILITIES','медицинские услуги':'MEDICAL',
 'обязательные пенсионные взносы работодателя':'TAXES','отчисления осмс':'TAXES',
 'социальные отчисления':'TAXES','социальный налог':'TAXES','прочие расходы':'OTHER',
 'расходы на корреспонденцию':'DELIVERY','расходы на наем жилого помещения':'LODGING','суточные':'LODGING',
 'расходы на питание':'MEALS','расходы проживание':'MEALS','расходы на проезд':'TRAVEL',
 'расходы по доставке':'DELIVERY','расходы по обучению':'TRAINING','расходы по перевозке':'TRANSPORT',
 'расходы по сиз':'PPE','ремонт авто':'REPAIR','расходы по ремонту оборудования':'REPAIR',
 'ремонт оборудования':'REPAIR','расходы по ремонту':'REPAIR','расходы по ремонту авто':'REPAIR',
 'расходы по экологии':'UTIL','экология':'UTIL','расходы по утилизации':'UTIL','утилизация':'UTIL',
 'расходы на билеты':'TRAVEL','билеты':'TRAVEL','авиабилеты':'TRAVEL','аренда оборудования':'RENT_TECH','геодезические услуги':'OTHER',
 'технические услуги':'OTHER','расходы по технической диагностике':'OTHER',
 'техническая диагностика':'OTHER','тех диагностика':'OTHER',
 'сертификационные услуги':'OTHER','сертификац':'OTHER',
 'себестоимость реализованной продукции и оказанных услуг':'COGS','содержание офиса':'SITE_MAINT',
 'списание материалов':'MATERIALS','страхование':'INSURANCE',
}
KEYWORD_RULES=[
 (('социальн','налог'),'TAXES'),(('соц','налог'),'TAXES'),
 (('ремонт','оборудован'),'REPAIR'),(('ремонт','авто'),'REPAIR'),
 (('аренда','оборудован'),'RENT_TECH'),(('аренда','спецтехн'),'RENT_TECH'),
 (('аренда','транспорт'),'RENT_TECH'),(('аренда','техник'),'RENT_TECH'),(('аренда','офис'),'RENT_OFFICE'),
 (('наем','жил'),'LODGING'),(('содержан','офис'),'SITE_MAINT'),(('содержан','сайт'),'SITE_MAINT'),
 (('услуги','связ'),'INTERNET'),(('техническ','диагностик'),'OTHER'),
 (('технические','услуг'),'OTHER'),
 (('амортизац',),'AMORT'),(('зарплат',),'SALARY'),(('оклад',),'SALARY'),(('заработн',),'SALARY'),
 (('геодез',),'OTHER'),(('подряд',),'SUBCONTRACT'),(('сертификац',),'OTHER'),
 (('опв',),'TAXES'),(('осмс',),'TAXES'),(('восмс',),'TAXES'),(('пенсионн',),'TAXES'),
 (('отчислен',),'TAXES'),(('налог',),'TAXES'),(('взнос',),'TAXES'),
 (('гсм',),'FUEL'),(('топлив',),'FUEL'),(('gps',),'GPS'),
 (('интернет',),'INTERNET'),(('связ',),'INTERNET'),
 (('коммунальн',),'UTILITIES'),(('электроэнерг',),'UTILITIES'),(('ком.услуг',),'UTILITIES'),
 (('медицин',),'MEDICAL'),(('медосмотр',),'MEDICAL'),(('мед осмотр',),'MEDICAL'),
 (('эколог',),'UTIL'),(('утилизац',),'UTIL'),
 (('билет',),'TRAVEL'),(('проезд',),'TRAVEL'),(('командиров',),'TRAVEL'),
 (('суточн',),'LODGING'),(('питан',),'MEALS'),(('прожив',),'MEALS'),
 (('доставк',),'DELIVERY'),(('корреспонд',),'DELIVERY'),(('обучен',),'TRAINING'),
 (('перевозк',),'TRANSPORT'),(('сиз',),'PPE'),(('ремонт',),'REPAIR'),(('аренда',),'RENT_TECH'),
 (('страхов',),'INSURANCE'),(('списан',),'MATERIALS'),(('материал',),'MATERIALS'),
 (('себестоимост',),'COGS'),(('канц',),'OTHER'),
 (('банк',),'BANK'),(('комисси',),'BANK'),(('подписк',),'SUBSCRIPTION'),
 (('лизинг',),'LEASING'),(('аудит',),'AUDIT'),(('консультац',),'AUDIT'),
]

ADMIN_LABEL={
 'SALARY':'Расходы по заработной плате','TAXES':'Налоги','FUEL':'ГСМ','BANK':'Банковская комиссия',
 'AMORT':'Амортизация ФА','SUBSCRIPTION':'Подписка','AUDIT':'Аудиторские/ консультационные услуги',
 'INTERNET':'Услуги связи+интернет','MEDICAL':'Мед осмотр','INSURANCE':'Расходы по страховке',
 'TRAVEL':'Командировочные расходы(проезд+проживание+суточные)',
 'LODGING':'Командировочные расходы(проезд+проживание+суточные)',
 'MEALS':'Командировочные расходы(проезд+проживание+суточные)','SITE_MAINT':'Содержание офиса',
 'TRAINING':'Обучение сотрудников','DELIVERY':'Услуги по доставке корресп.и др.',
 'LEASING':'Вознаграждения по лизингу','OTHER':'Прочие расходы',
}

def resolve_concept(nit):
    """nit — нормализованное название статьи. Точное совпадение -> ключевые слова."""
    if nit in OPEX_CONCEPT: return OPEX_CONCEPT[nit]
    for kws,concept in KEYWORD_RULES:
        if all(k in nit for k in kws): return concept
    return None

OPEX_LABEL={
 'TCO':{'GPS':'GPS','AMORT':'Амортизация','RENT_OFFICE':'Аренда офиса','RENT_TECH':'Аренда техники','FUEL':'ГСМ',
   'SALARY':'Заработная плата','INTERNET':'Интернет+ связь','MEDICAL':'Мед.обслуживание','TAXES':'Налоги',
   'OTHER':'Прочие услуги сторонних организаций','LODGING':'Расходны на наем жилья и Суточные',
   'MEALS':'Расходы на питание и проживание','TRAVEL':'Расходы на проезд','DELIVERY':'Расходы по доставке',
   'TRAINING':'Расходы по обучению','TRANSPORT':'Расходы по перевозке','PPE':'Расходы по СИЗ','UTIL':'Расходы на утилизацию',
   'REPAIR':'Расходы на ремонт авто/ оборудования','COGS':'Услуги сторонних/ подрядных организаций',
   'SITE_MAINT':'Содержание на сайте офиса+ контейнера','MATERIALS':'Списанные материалы','INSURANCE':'Страхование',
   'UTILITIES':'Содержание на сайте офиса+ контейнера','BASE_MAINT':'Прочие услуги сторонних организаций','SUBCONTRACT':'Услуги сторонних/ подрядных организаций'},
 'KPO':{'GPS':'GPS','AMORT':'Амортизация','RENT_OFFICE':'Аренда офиса','RENT_TECH':'Аренда техники','FUEL':'ГСМ',
   'SALARY':'Заработная плата','INTERNET':'Интернет+связь','MEDICAL':'Мед.обслуживание','TAXES':'Налоги',
   'OTHER':'Прочие услуги','LODGING':'Расходны на наем жилья и Суточные','MEALS':'Расходы на питание и проживание',
   'TRAVEL':'Расходы на проезд','DELIVERY':'Расходы по доставке','TRAINING':'Расходы по обучению',
   'TRANSPORT':'Расходы по перевозке','PPE':'Расходы по СИЗ','UTIL':'Расходы на утилизацию',
   'REPAIR':'Расходы на ремонт авто/ оборудования','COGS':'Услуги сторонних/ подрядных организаций',
   'SITE_MAINT':'Содержание на сайте офиса+ контейнера','MATERIALS':'Списанные материалы','INSURANCE':'Страхование',
   'UTILITIES':'Содержание на сайте офиса+ контейнера','BASE_MAINT':'Прочие услуги','SUBCONTRACT':'Услуги сторонних/ подрядных организаций'},
 'CONCRETE':{'GPS':'GPS','RENT_TECH':'Аренда техники','BASE_MAINT':'База содержание','FUEL':'ГСМ',
   'SALARY':'Заработная плата','INTERNET':'Интернет+связь','UTILITIES':'Коммунальные услуги','TAXES':'Налоги',
   'OTHER':'Прочие расходы','LODGING':'Расходны на наем жилья (Эркан)  и Суточные','TRAVEL':'Расходы на проезд',
   'TRANSPORT':'Расходы по перевозке','PPE':'Расходы по СИЗ','UTIL':'Расходы по экологии',
   'REPAIR':'Расходы по ремонту авто/ оборудования','COGS':'Себестоимость товарного бетона',
   'MATERIALS':'Списанные материалы','INSURANCE':'Страхование','MEALS':'Прочие расходы','DELIVERY':'Прочие расходы',
   'TRAINING':'Прочие расходы','MEDICAL':'Прочие расходы','RENT_OFFICE':'Прочие расходы','SITE_MAINT':'Прочие расходы',
   'AMORT':'Амортизация','SUBCONTRACT':'Прочие расходы'},
}
INCOME_LABEL={
 'TCO':{'INC_SALES':'Доход от реализации','INC_FIN':'__OTHER_INCOME__','INC_OTHER':'__OTHER_INCOME__'},
 'KPO':{'INC_SALES':'Доход от реализации','INC_FIN':'__OTHER_INCOME__','INC_OTHER':'__OTHER_INCOME__'},
 'CONCRETE':{'INC_CONCRETE':'Доход от реализации бетона','INC_SALES':'Доход от реализации бетона','INC_RENT':'Доход от сдачи в аренду',
   'INC_FIN':'Прочий доход','INC_SUBSIDY':'Прочие доходы (субсидии Даму)','INC_OTHER':'Прочий доход'},
}
ADMIN_CONCEPT={
 'зарплата':'Расходы по заработной плате','обязательные пенсионные взносы работодателя':'Налоги',
 'отчисления осмс':'Налоги','социальные отчисления':'Налоги','социальный налог':'Налоги','гсм':'ГСМ',
 'комиссия банка':'Банковская комиссия','амортизация фа':'Амортизация ФА','подписка':'Подписка',
 'интернет':'Услуги связи+интернет','услуги связи':'Услуги связи+интернет','медицинские услуги':'Мед осмотр',
 'страхование':'Расходы по страховке','суточные':'Командировочные расходы(проезд+проживание+суточные)',
 'расходы на наем жилого помещения':'Командировочные расходы(проезд+проживание+суточные)',
 'расходы на проезд':'Командировочные расходы(проезд+проживание+суточные)','содержание офиса':'Содержание офиса',
 'обучение сотрудников':'Обучение сотрудников','расходы по обучению':'Обучение сотрудников','расходы на билеты':'Командировочные расходы(проезд+проживание+суточные)','билеты':'Командировочные расходы(проезд+проживание+суточные)','канц товары':'Прочие расходы','прочие расходы':'Прочие расходы',
 'расходы на корреспонденцию':'Услуги по доставке корресп.и др.',
}
DIV_KEY={'ТШО':'TCO','КПО':'KPO','Бетон':'CONCRETE'}

def block_index(ws,start,end):
    idx={}
    for r in range(start,end+1):
        v=ws.cell(row=r,column=2).value
        if v is not None: idx[str(v).strip()]=r
    return idx

def other_income_row(ws):
    for r in range(28,40):
        v=ws.cell(row=r,column=2).value
        if v and 'прочие доходы' in str(v).lower(): return r
    return 31

def fill_report(template_path,leaves,out_path,month_col='C'):
    wb=openpyxl.load_workbook(template_path)
    tco,kpo,con,alls=wb['P&L 2026 TCO'],wb['P&L KPO'],wb['P&L 2026 concrete'],wb['P&L 2026 all']
    SH={'TCO':tco,'KPO':kpo,'CONCRETE':con}
    OPX={'TCO':block_index(tco,7,30),'KPO':block_index(kpo,7,30),'CONCRETE':block_index(con,9,28)}
    INC={'TCO':block_index(tco,3,5),'KPO':block_index(kpo,3,5),'CONCRETE':block_index(con,3,6)}
    OTH={'TCO':other_income_row(tco),'KPO':other_income_row(kpo)}
    ADM=block_index(con,31,46)
    def _last_row(ws,label,lo,hi):
        f=None
        for r in range(lo,hi+1):
            v=ws.cell(row=r,column=2).value
            if v and str(v).strip()==label: f=r
        return f
    BELOWPR={'TCO':_last_row(tco,'Прочие расходы',7,45),'KPO':_last_row(kpo,'Прочие расходы',7,45),'CONCRETE':_last_row(con,'Прочие расходы',9,55)}
    acc=defaultdict(float); flags=[]; log=[]
    def add(ws,row,amt): acc[(ws.title,row)]+=amt
    def by_label(pool,label):
        if not label: return None
        lab=label.strip()
        for k,r in pool.items():
            if k.strip()==lab: return r
        return None
    def put(dk,label,amt,is_income=False):
        ws=SH[dk]
        if label=='__OTHER_INCOME__': add(ws,OTH[dk],amt); return ws.cell(row=OTH[dk],column=2).value
        if label=='__CONCRETE_ADMIN_AMORT__': add(con,ADM['Амортизация ФА'],amt); return 'Амортизация ФА (Бетон, адм.)'
        pool=INC[dk] if is_income else OPX[dk]
        r=by_label(pool,label)
        if r: add(ws,r,amt); return label.strip()
        return None

    for (account,item,kind,div,amount) in leaves:
        account=account or ''
        if kind=='control': continue
        nit=norm(item)
        if kind=='income':
            _,concept=INCOME_MAP.get((account,nit),(None,'INC_OTHER'))
            d=div if div in DIV_KEY else 'Бетон'   # подразделение из выгрузки, не из словаря
            dk=DIV_KEY[d]; lab=put(dk,INCOME_LABEL[dk].get(concept,'__OTHER_INCOME__'),amount,True)
            log.append((account,item,d,'доход',SH[dk].title,lab,amount)); continue
        # КПН (77xx) считается формулой шаблона — пропускаем
        if account[:2]=='77':
            log.append((account,item,div or '—','кпн(пропуск)','—','—',amount)); continue
        # Финансирование 73xx -> операционная строка «Расходы на финансирование» по подразделению
        if account[:2]=='73':
            if div in DIV_KEY: dk=DIV_KEY[div]
            else: dk='CONCRETE'; flags.append((item,div or '—',amount,'финансирование без подразделения → concrete'))
            lab=put(dk,'Расходы на финансирование',amount)
            log.append((account,item,div or '—','расход',SH[dk].title,lab,amount)); continue
        # 7400 (курсовые 7430, обмен валюты 7480 и пр.) -> строка «Прочие расходы» (ниже операц.)
        if account[:2]=='74':
            if div in DIV_KEY: dk=DIV_KEY[div]
            else: dk='CONCRETE'; flags.append((item,div or '—',amount,'курсовые/прочие без подразделения → concrete'))
            add(SH[dk], BELOWPR[dk], amount)
            log.append((account,item,div or '—','прочие расх.',SH[dk].title,'Прочие расходы',amount)); continue
        if div=='АУП':
            concept=resolve_concept(nit)
            label=ADMIN_LABEL.get(concept) if concept else None
            if label and label in ADM:
                add(con,ADM[label],amount); log.append((account,item,div,'админ',con.title,label,amount))
            else:
                if amount>=THRESHOLD and concept is None: flags.append((item,div,amount,'новая АУП-статья ≥300к'))
                add(con,ADM['Прочие расходы'],amount); log.append((account,item,div,'админ',con.title,'Прочие расходы',amount))
            continue
        if div not in DIV_KEY:
            flags.append((item, div or '—', amount, 'статья без подразделения → concrete, проверьте'))
            div='Бетон'
        dk=DIV_KEY[div]; concept=resolve_concept(nit)
        if concept:
            lab=put(dk,OPEX_LABEL[dk].get(concept),amount)
            if lab is None:
                if amount>=THRESHOLD: flags.append((item,div,amount,f'≥300к, нет строки на {SH[dk].title}'))
                lab=put(dk,OPEX_LABEL[dk].get('OTHER'),amount)
            log.append((account,item,div,'расход',SH[dk].title,lab,amount))
        else:
            if amount>=THRESHOLD: flags.append((item,div,amount,f'новая статья ≥300к на {SH[dk].title}'))
            lab=put(dk,OPEX_LABEL[dk].get('OTHER'),amount)
            log.append((account,item,div,'расход(нов)',SH[dk].title,lab,amount))

    name2ws={ws.title:ws for ws in (tco,kpo,con)}
    for (sn,row),val in acc.items(): name2ws[sn][f'{month_col}{row}']=round(val,2)
    wire_all_sheet(alls,tco,kpo,con)
    # контроль сверки с ОСВ: 5610 и 7710 в блок на листе all
    ctrl={lf[0]:lf[4] for lf in leaves if lf[2]=='control'}
    def _row_all(label):
        for r in range(55,75):
            if (alls.cell(row=r,column=2).value or '').strip()==label: return r
    if ctrl:
        r5=_row_all('ОСВ 1С (сч. 5610)'); r7=_row_all('КПН (сч. 7710)')
        if r5 and '5610' in ctrl: alls[f'{month_col}{r5}']=round(ctrl['5610'],2)
        if r7 and '7710' in ctrl: alls[f'{month_col}{r7}']=round(ctrl['7710'],2)
    wb.save(out_path)
    return acc,flags,log

def wire_all_sheet(alls,tco,kpo,con):
    def find(ws,text,a,b):
        for r in range(a,b+1):
            v=ws.cell(row=r,column=2).value
            if v and str(v).strip()==text: return r
        return None
    r5=find(alls,'Прочий доход',3,8); toi=other_income_row(tco); koi=other_income_row(kpo)
    if r5:
        for col in 'CDEFGHIJKLMN':
            alls[f'{col}{r5}']=f"='P&L 2026 concrete'!{col}5+'P&L 2026 TCO'!{col}{toi}+'P&L KPO'!{col}{koi}"
    a0=find(alls,'Расходы по заработной плате',35,52); c0=find(con,'Расходы по заработной плате',28,45)
    if a0 and c0:
        for i in range(16):
            for col in 'CDEFGHIJKLMN':
                alls[f'{col}{a0+i}']=f"='P&L 2026 concrete'!{col}{c0+i}"
    # нижне-операционная «Прочие расходы»: all = сумма дивизионов (по факту строк, не по номеру)
    def lastrow(ws,text,a,b):
        fr=None
        for r in range(a,b+1):
            v=ws.cell(row=r,column=2).value
            if v and str(v).strip()==text: fr=r
        return fr
    apr=lastrow(alls,'Прочие расходы',50,60); tpr=lastrow(tco,'Прочие расходы',7,45)
    kpr=lastrow(kpo,'Прочие расходы',7,45); cpr=lastrow(con,'Прочие расходы',9,55)
    if apr and tpr and kpr and cpr:
        for col in 'CDEFGHIJKLMN':
            alls[f'{col}{apr}']=f"='P&L 2026 TCO'!{col}{tpr}+'P&L KPO'!{col}{kpr}+'P&L 2026 concrete'!{col}{cpr}"
    # унификация: переносим корректные формулы столбца C во все месяцы D..N
    from openpyxl.formula.translate import Translator
    from openpyxl.utils import get_column_letter as _gcl
    for r in range(3,66):
        cval=alls.cell(row=r,column=3).value
        if isinstance(cval,str) and cval.startswith('='):
            for col in range(4,15):
                L=_gcl(col)
                try: alls.cell(row=r,column=col).value=Translator(cval,origin=f'C{r}').translate_formula(f'{L}{r}')
                except Exception: pass

if __name__=='__main__':
    leaves=parse_1c('/mnt/user-data/uploads/январь_2026.xlsx')
    acc,flags,log=fill_report('/mnt/user-data/uploads/GRD_2022-2026_1С_.xlsx',leaves,'/home/claude/GRD_январь_заполнен.xlsx')
    print('cells:',len(acc),'flags:',len(flags))
    for f in flags: print('  FLAG:',f)
